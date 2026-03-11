"""
Demo runner for MobileAgent v3 with Excel output.
Supports GUI-Owl (vLLM) and GPT-4o (OpenAI API).

Uses GUIOwlWrapper for GUI-Owl models (matching the paper) and
GPT4Wrapper for GPT-4o.
"""
import os
import sys
import uuid
import json
import time
import argparse
from PIL import Image
from datetime import datetime

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from utils.mobile_agent_e import (
    InfoPool,
    Manager,
    Executor,
    Notetaker,
    ActionReflector,
    INPUT_KNOW,
)
import utils.controller as controller
from utils.call_mobile_agent_e import GUIOwlWrapper, GPT4Wrapper


HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
PENDING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")

HEADERS = [
    "Step",
    "Agent",
    "Thought",
    "Plan / Action",
    "Description",
    "Outcome",
    "Screenshot (Before)",
    "Screenshot (After)",
    "Timestamp",
]

COL_WIDTHS = [6, 12, 50, 50, 40, 14, 22, 22, 20]


def _is_gui_owl(model_name: str) -> bool:
    name = model_name.lower()
    return "gui-owl" in name or "gui_owl" in name


def init_workbook(instruction: str, model_name: str = "mPLUG/GUI-Owl-7B"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Steps"

    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"Test Case: {instruction}"
    title_cell.font = Font(bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:I2")
    meta = ws["A2"]
    meta.value = f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  Model: {model_name}  |  Status: RUNNING"
    meta.font = Font(size=10, italic=True)
    meta.alignment = Alignment(horizontal="center")

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for col_idx, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    return wb, ws


def add_step_row(ws, row, step, agent, thought, action_or_plan, description, outcome,
                 screenshot_before=None, screenshot_after=None):
    data = [step, agent, thought, action_or_plan, description, outcome, "", "", datetime.now().strftime("%H:%M:%S")]
    for col_idx, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.alignment = WRAP_ALIGN
        cell.border = THIN_BORDER

    if outcome:
        outcome_cell = ws.cell(row=row, column=6)
        if "A" in outcome or "Success" in outcome or "Finished" in outcome:
            outcome_cell.fill = SUCCESS_FILL
        elif "B" in outcome or "C" in outcome or "Fail" in outcome:
            outcome_cell.fill = FAIL_FILL
        else:
            outcome_cell.fill = PENDING_FILL

    ws.row_dimensions[row].height = 120

    def _embed_image(col, img_path):
        if img_path and os.path.exists(img_path):
            try:
                thumb_path = img_path.replace(".png", "_thumb.png")
                img = Image.open(img_path)
                img.thumbnail((200, 400))
                img.save(thumb_path)
                xl_img = XlImage(thumb_path)
                cell_ref = f"{get_column_letter(col)}{row}"
                ws.add_image(xl_img, cell_ref)
            except Exception as e:
                ws.cell(row=row, column=col, value=f"[img error: {e}]")

    _embed_image(7, screenshot_before)
    _embed_image(8, screenshot_after)


def run_demo(adb_path, api_key, base_url, model, instruction, add_info, max_step=25,
             log_path="./logs", excel_path=None, coor_type="abs"):
    from utils.android_controller import AndroidController
    ctrl = AndroidController(adb_path)

    if not os.path.exists(log_path):
        os.makedirs(log_path)

    now = datetime.now()
    time_str = now.strftime("%Y%m%d_%H%M%S")
    save_path = os.path.join(log_path, f"{time_str}_{instruction[:20].replace(' ', '_')}")
    os.makedirs(save_path, exist_ok=True)
    image_save_path = os.path.join(save_path, "images")
    os.makedirs(image_save_path, exist_ok=True)

    if excel_path is None:
        excel_path = os.path.join(save_path, "test_report.xlsx")

    wb, ws = init_workbook(instruction, model_name=model)
    excel_row = 4

    info_pool = InfoPool(
        additional_knowledge_manager=add_info,
        additional_knowledge_executor=INPUT_KNOW,
        err_to_manager_thresh=2,
    )

    if _is_gui_owl(model):
        vllm = GUIOwlWrapper(api_key, base_url, model)
        print(f"  Wrapper : GUIOwlWrapper (paper)")
    else:
        vllm = GPT4Wrapper(api_key, base_url, model)
        print(f"  Wrapper : GPT4Wrapper")

    manager = Manager()
    executor = Executor()
    notetaker = Notetaker()
    action_reflector = ActionReflector()
    message_manager, message_operator, message_reflector, message_notekeeper = None, None, None, None
    info_pool.instruction = instruction

    final_status = "INCOMPLETE"

    for step in range(max_step):
        print(f"\n{'='*60}")
        print(f"  STEP {step + 1}")
        print(f"{'='*60}")

        if step == 0:
            current_time = datetime.now()
            formatted_time = current_time.strftime(
                f'%Y-%m-%d-{current_time.hour * 3600 + current_time.minute * 60 + current_time.second}-{uuid.uuid4().hex[:8]}'
            )
            local_image_dir = os.path.join(image_save_path, f"screenshot_{formatted_time}.png")
        else:
            local_image_dir = local_image_dir2

        for _ in range(5):
            if not ctrl.get_screenshot(local_image_dir):
                print("  Get screenshot failed, retry...")
                time.sleep(5)
            else:
                break

        width, height = Image.open(local_image_dir).size

        info_pool.error_flag_plan = False
        err_to_manager_thresh = info_pool.err_to_manager_thresh
        if len(info_pool.action_outcomes) >= err_to_manager_thresh:
            latest_outcomes = info_pool.action_outcomes[-err_to_manager_thresh:]
            count = sum(1 for o in latest_outcomes if o in ("B", "C"))
            if count == err_to_manager_thresh:
                info_pool.error_flag_plan = True

        skip_manager = False
        if not info_pool.error_flag_plan and len(info_pool.action_history) > 0:
            if info_pool.action_history[-1]["action"] == "invalid":
                skip_manager = True

        # --- Manager ---
        if not skip_manager:
            print("\n  [Manager] Planning...")
            prompt_planning = manager.get_prompt(info_pool)
            output_planning, message_manager, raw_response = vllm.predict_mm(prompt_planning, [local_image_dir])

            msg_save_path = os.path.join(save_path, f"step_{step + 1}")
            os.makedirs(msg_save_path, exist_ok=True)
            with open(os.path.join(msg_save_path, "manager.json"), "w", encoding="utf-8") as f:
                json.dump({"name": "manager", "messages": message_manager, "response": output_planning, "step_id": step + 1}, f, ensure_ascii=False, indent=4)

            parsed_planning = manager.parse_response(output_planning)
            info_pool.completed_plan = parsed_planning["completed_subgoal"]
            info_pool.plan = parsed_planning["plan"]
            if not raw_response:
                raise RuntimeError("Error calling LLM in planning phase.")

            print(f"  Completed: {info_pool.completed_plan}")
            print(f"  Thought: {parsed_planning['thought'][:120]}...")
            print(f"  Plan: {info_pool.plan}")

            add_step_row(ws, excel_row, step + 1, "Manager",
                         parsed_planning["thought"][:500],
                         info_pool.plan[:500],
                         info_pool.completed_plan[:300],
                         "Planning",
                         screenshot_before=local_image_dir)
            excel_row += 1
            wb.save(excel_path)

        if "Finished" in info_pool.plan.strip() and len(info_pool.plan.strip()) < 15:
            print("\n  >>> INSTRUCTION FINISHED <<<")
            final_status = "COMPLETED"

            add_step_row(ws, excel_row, step + 1, "System", "Task completed", "Finished", instruction, "Finished",
                         screenshot_before=local_image_dir)
            excel_row += 1

            task_result_path = os.path.join(save_path, "task_result.json")
            task_result_data = {
                "goal": instruction,
                "finish_dtime": datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f"),
                "hit_step_limit": 0.0,
            }
            with open(task_result_path, "w", encoding="utf-8") as f:
                json.dump(task_result_data, f, ensure_ascii=False, indent=4)
            break

        # --- Executor ---
        print("\n  [Executor] Deciding action...")
        if not os.path.exists(os.path.join(save_path, f"step_{step + 1}")):
            os.makedirs(os.path.join(save_path, f"step_{step + 1}"), exist_ok=True)

        prompt_action = executor.get_prompt(info_pool)
        output_action, message_operator, raw_response = vllm.predict_mm(prompt_action, [local_image_dir])

        if not raw_response:
            raise RuntimeError("Error calling LLM in operator phase.")
        parsed_action = executor.parse_response(output_action)
        action_thought = parsed_action["thought"]
        action_object_str = parsed_action["action"]
        action_description = parsed_action["description"]

        info_pool.last_action_thought = action_thought
        info_pool.last_summary = action_description

        if not action_thought or not action_object_str:
            print("  Action output invalid format.")
            info_pool.last_action = {"action": "invalid"}
            info_pool.action_history.append({"action": "invalid"})
            info_pool.summary_history.append(action_description)
            info_pool.action_outcomes.append("C")
            info_pool.error_descriptions.append("invalid action format, do nothing.")

            add_step_row(ws, excel_row, step + 1, "Executor", action_thought or "N/A",
                         action_object_str or "N/A", action_description or "invalid", "C - Invalid",
                         screenshot_before=local_image_dir)
            excel_row += 1
            wb.save(excel_path)
            continue

        action_object_str = action_object_str.replace("```", "").replace("json", "").strip()
        print(f"  Thought: {action_thought[:120]}...")
        print(f"  Action: {action_object_str}")
        print(f"  Description: {action_description}")

        try:
            action_object = json.loads(action_object_str)

            if action_object["action"] == "answer":
                answer_content = action_object["text"]
                print(f"\n  >>> ANSWER: {answer_content} <<<")
                final_status = "COMPLETED"

                msg_save_path = os.path.join(save_path, f"step_{step + 1}")
                with open(os.path.join(msg_save_path, "operator.json"), "w", encoding="utf-8") as f:
                    json.dump({"name": "operator", "messages": message_operator, "response": output_action, "step_id": step + 1}, f, ensure_ascii=False, indent=4)

                add_step_row(ws, excel_row, step + 1, "Executor", action_thought[:500],
                             f"answer: {answer_content}", action_description, "Finished",
                             screenshot_before=local_image_dir)
                excel_row += 1

                task_result_data = {
                    "goal": instruction,
                    "finish_dtime": datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f"),
                    "hit_step_limit": 0.0,
                }
                with open(os.path.join(save_path, "task_result.json"), "w", encoding="utf-8") as f:
                    json.dump(task_result_data, f, ensure_ascii=False, indent=4)
                break

            if coor_type != "abs":
                if "coordinate" in action_object:
                    action_object["coordinate"] = [
                        int(action_object["coordinate"][0] / 1000 * width),
                        int(action_object["coordinate"][1] / 1000 * height),
                    ]
                if "coordinate2" in action_object:
                    action_object["coordinate2"] = [
                        int(action_object["coordinate2"][0] / 1000 * width),
                        int(action_object["coordinate2"][1] / 1000 * height),
                    ]

            if action_object["action"] == "click":
                ctrl.tap(action_object["coordinate"][0], action_object["coordinate"][1])
            elif action_object["action"] == "swipe":
                ctrl.slide(action_object["coordinate"][0], action_object["coordinate"][1],
                           action_object["coordinate2"][0], action_object["coordinate2"][1])
            elif action_object["action"] == "type":
                ctrl.type(action_object["text"])
            elif action_object["action"] == "system_button":
                if action_object["button"] == "Back":
                    ctrl.back()
                elif action_object["button"] == "Home":
                    ctrl.home()

        except Exception as e:
            print(f"  Action execution error: {e}")
            info_pool.last_action = {"action": "invalid"}
            info_pool.action_history.append({"action": "invalid"})
            info_pool.summary_history.append(action_description)
            info_pool.action_outcomes.append("C")
            info_pool.error_descriptions.append("invalid action format, do nothing.")
            local_image_dir2 = local_image_dir

            add_step_row(ws, excel_row, step + 1, "Executor", action_thought[:500],
                         action_object_str[:500], f"ERROR: {e}", "C - Error",
                         screenshot_before=local_image_dir)
            excel_row += 1
            wb.save(excel_path)
            continue

        msg_save_path = os.path.join(save_path, f"step_{step + 1}")
        os.makedirs(msg_save_path, exist_ok=True)
        with open(os.path.join(msg_save_path, "operator.json"), "w", encoding="utf-8") as f:
            json.dump({"name": "operator", "messages": message_operator, "response": output_action, "step_id": step + 1}, f, ensure_ascii=False, indent=4)

        info_pool.last_action = json.loads(action_object_str)

        if step == 0:
            time.sleep(8)
        time.sleep(2)

        current_time = datetime.now()
        formatted_time = current_time.strftime(
            f'%Y-%m-%d-{current_time.hour * 3600 + current_time.minute * 60 + current_time.second}-{uuid.uuid4().hex[:8]}'
        )
        local_image_dir2 = os.path.join(image_save_path, f"screenshot_{formatted_time}.png")

        for _ in range(5):
            if not ctrl.get_screenshot(local_image_dir2):
                print("  Get screenshot failed, retry...")
                time.sleep(5)
            else:
                break

        # --- Action Reflector ---
        print("\n  [Reflector] Evaluating action...")
        prompt_reflect = action_reflector.get_prompt(info_pool)
        output_reflect, message_reflector, raw_response = vllm.predict_mm(
            prompt_reflect, [local_image_dir, local_image_dir2]
        )

        with open(os.path.join(msg_save_path, "reflector.json"), "w", encoding="utf-8") as f:
            json.dump({"name": "reflector", "messages": message_reflector, "response": output_reflect, "step_id": step + 1}, f, ensure_ascii=False, indent=4)

        parsed_reflect = action_reflector.parse_response(output_reflect)
        outcome = parsed_reflect["outcome"]
        error_description = parsed_reflect["error_description"]

        if "A" in outcome:
            action_outcome = "A"
        elif "B" in outcome:
            action_outcome = "B"
        elif "C" in outcome:
            action_outcome = "C"
        else:
            action_outcome = "C"

        outcome_label = {
            "A": "A - Success",
            "B": "B - Wrong Page",
            "C": "C - No Change",
        }.get(action_outcome, action_outcome)

        print(f"  Outcome: {outcome_label}")
        if error_description and error_description != "None":
            print(f"  Error: {error_description}")

        info_pool.action_history.append(json.loads(action_object_str))
        info_pool.summary_history.append(action_description)
        info_pool.action_outcomes.append(action_outcome)
        info_pool.error_descriptions.append(error_description)
        info_pool.progress_status = info_pool.completed_plan

        add_step_row(ws, excel_row, step + 1, "Executor",
                     action_thought[:500],
                     action_object_str[:500],
                     action_description[:300],
                     outcome_label,
                     screenshot_before=local_image_dir,
                     screenshot_after=local_image_dir2)
        excel_row += 1
        wb.save(excel_path)
        print(f"  [Excel] Saved step {step + 1} to {excel_path}")

    # Final summary row
    ws.merge_cells(f"A{excel_row}:I{excel_row}")
    summary_cell = ws.cell(row=excel_row, column=1)
    summary_cell.value = f"FINAL STATUS: {final_status}  |  Total Steps: {excel_row - 4}  |  Ended: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    summary_cell.font = Font(bold=True, size=12, color="1F4E79")
    summary_cell.alignment = Alignment(horizontal="center")
    if final_status == "COMPLETED":
        summary_cell.fill = SUCCESS_FILL
    else:
        summary_cell.fill = FAIL_FILL

    ws["A2"].value = f"Started: {now.strftime('%Y-%m-%d %H:%M:%S')}  |  Model: {model}  |  Status: {final_status}"

    wb.save(excel_path)
    print(f"\n{'='*60}")
    print(f"  DONE! Status: {final_status}")
    print(f"  Excel report: {excel_path}")
    print(f"  Logs: {save_path}")
    print(f"{'='*60}")
    return excel_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="MobileAgent v3 Demo with Excel Output")
    parser.add_argument("--adb_path", type=str, required=True)
    parser.add_argument("--api_key", type=str, default=None,
                        help="API key. Falls back to OPENAI_API_KEY env var if not set.")
    parser.add_argument("--base_url", type=str, default=None,
                        help="Base URL. Auto-detected from --model if not set.")
    parser.add_argument("--model", type=str, default="mPLUG/GUI-Owl-7B",
                        help="'mPLUG/GUI-Owl-7B' -> GUIOwlWrapper, 'gpt-4o' -> GPT4Wrapper")
    parser.add_argument("--instruction", type=str, required=True)
    parser.add_argument("--add_info", type=str, default="")
    parser.add_argument("--max_step", type=int, default=25)
    parser.add_argument("--coor_type", type=str, default="abs",
                        help="'abs' for absolute pixel coords, 'qwen-vl' for 0-1000 range")
    parser.add_argument("--excel_path", type=str, default=None)
    args = parser.parse_args()

    api_key = args.api_key or os.environ.get("OPENAI_API_KEY", "")
    if not api_key:
        parser.error("Provide --api_key or set the OPENAI_API_KEY environment variable.")

    if args.base_url:
        base_url = args.base_url
    elif _is_gui_owl(args.model):
        base_url = "https://male-richards-chester-his.trycloudflare.com/v1"
    else:
        base_url = "https://api.openai.com/v1"

    wrapper = "GUIOwlWrapper" if _is_gui_owl(args.model) else "GPT4Wrapper"
    print(f"Model   : {args.model}")
    print(f"Wrapper : {wrapper}")
    print(f"Base URL: {base_url}")

    run_demo(
        adb_path=args.adb_path,
        api_key=api_key,
        base_url=base_url,
        model=args.model,
        instruction=args.instruction,
        add_info=args.add_info,
        max_step=args.max_step,
        excel_path=args.excel_path,
        coor_type=args.coor_type,
    )
